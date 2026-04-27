"""sym-ui issue_log.xlsx 생성 스크립트.

글로벌 CLAUDE.md 의 12컬럼 양식을 따른다.
한 번 만들면 이후 이슈는 openpyxl 로 append 한다.
"""
from __future__ import annotations
import sys
from pathlib import Path

sys.stdout.reconfigure(encoding="utf-8")

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

OUT = Path(__file__).resolve().parent.parent / "issue_log.xlsx"

HEADERS = [
    "이슈번호", "발견일", "유형", "제목", "사전 조건",
    "재현 방법 / 개선 사항", "원인 및 수정 / 개선 사유",
    "영향 범위", "발생 버전", "수정 버전", "상태", "심각도",
]

# (이슈번호, 발견일, 유형, 제목, 사전조건, 재현/개선, 원인/사유, 영향범위, 발생버전, 수정버전, 상태, 심각도)
ROWS = [
    # === Critical 4 ===
    (1, "2026-04-27", "bug",
     "다크 토큰 시스템 부재로 26개 컴포넌트가 dark:[#hex] 하드코딩 우회",
     "다크모드 사용",
     "1. 컴포넌트 그렙 시 dark:[#1e222d], dark:[#2a2d3e] 등 하드 hex 가 곳곳에 산재\n2. 시맨틱 토큰 부재로 토큰 변경 시 26개 파일 모두 수정 필요",
     "원인: tokens/colors.ts 와 tailwind.preset.cjs 에 다크 전용 시맨틱 토큰이 없어 우회 패턴 발생.\n수정: shadcn 스타일 CSS 변수 18종 (background, surface, muted, border, primary 등) 을 :root / .dark 양쪽에 정의하고 26개 컴포넌트를 모두 마이그레이션.",
     "26개 UI 컴포넌트 + globals.css + tailwind.preset.cjs",
     "@sym/ui 0.0.0", "@sym/ui 0.1.0", "완료", "Critical"),

    (2, "2026-04-27", "bug",
     "다크 배경 절대 밝기 L<15% 위반 (글로벌 다크 팔레트 규칙)",
     "다크모드 사용",
     "1. globals.css 의 --background L=6%\n2. 컴포넌트의 #131722 L=10.4%, #1e222d L=14.7% 모두 L<15%\n3. 결과적으로 순수 검정처럼 보임",
     "원인: 초기 팔레트 설계 시 절대 밝기 규칙 미반영.\n수정: 다크 --background L=15%, --surface L=18%, --surface-elevated L=22%, --border L=28% 로 전면 재구성. 모두 L>=15 보장.",
     "전 컴포넌트 다크모드 외관",
     "@sym/ui 0.0.0", "@sym/ui 0.1.0", "완료", "Critical"),

    (3, "2026-04-27", "bug",
     "Dialog 기본 사용 시 aria-describedby 누락으로 Radix 경고 + a11y 위반",
     "DialogContent 를 DialogDescription 없이 사용",
     "1. DialogContent 안에 DialogDescription 없이 렌더\n2. 콘솔에 Missing Description or aria-describedby={undefined} 경고\n3. 화면 낭독기에 Dialog 컨텍스트 누락",
     "원인: DialogContent 가 aria-describedby 를 명시 전달하지 않아 Radix 가 false positive 경고.\n수정: aria-describedby prop 을 명시 전달, JSDoc 에 DialogDescription 권장 명시, 테스트도 DialogDescription 포함.",
     "Dialog 사용 모든 페이지의 a11y",
     "@sym/ui 0.0.0", "@sym/ui 0.1.0", "완료", "Critical"),

    (4, "2026-04-27", "bug",
     "Sheet 기본 사용 시 aria-describedby 누락 (Dialog 와 동일)",
     "SheetContent 를 SheetDescription 없이 사용",
     "1. SheetContent 안에 SheetDescription 없이 렌더\n2. Radix 동일 경고 + 화면낭독기 컨텍스트 누락",
     "원인: SheetContent 도 같은 패턴.\n수정: Dialog 와 동일 패턴 적용 + sheet.test.tsx 업데이트.",
     "Sheet 사용 모든 페이지의 a11y",
     "@sym/ui 0.0.0", "@sym/ui 0.1.0", "완료", "Critical"),

    # === Major 14 ===
    (5, "2026-04-27", "bug",
     "Progress: max=0 일 때 NaN% 렌더링",
     "Progress 컴포넌트에 max={0} 전달",
     "1. <Progress value={0} max={0} /> 렌더\n2. (value/max)*100 = NaN\n3. style transform: translateX(-NaN%) 적용",
     "원인: max 가 0/음수일 때 가드 누락.\n수정: safeMax 변수 도입, max>0 인지 검사 후 100 으로 보정. value 검증도 safeMax 기준.",
     "Progress 컴포넌트",
     "@sym/ui 0.0.0", "@sym/ui 0.1.0", "완료", "Major"),

    (6, "2026-04-27", "bug",
     "CLI add: prompts 취소 시 selected=undefined 로 런타임 크래시",
     "sym-ui add 무인자 실행 후 ESC 또는 Ctrl+C",
     "1. sym-ui add 실행\n2. 다중선택 프롬프트에서 ESC\n3. selected=undefined 가 components 에 대입\n4. for...of 에서 TypeError",
     "원인: prompts 결과 undefined 가드 부재.\n수정: Array.isArray(selected) 검사 후 빈 배열이면 친절 메시지 후 return.",
     "@sym/ui-cli add 명령",
     "@sym/ui-cli 0.0.0", "@sym/ui-cli 0.1.0", "완료", "Major"),

    (7, "2026-04-27", "bug",
     "CLI init: prompts 취소 시 componentsDir=undefined 로 크래시",
     "sym-ui init 실행 후 ESC",
     "1. sym-ui init 실행\n2. 컴포넌트 경로 입력 프롬프트에서 ESC\n3. componentsDir=undefined 로 path.join 호출\n4. TypeError",
     "원인: text 프롬프트 결과 undefined 가드 부재.\n수정: typeof string 이고 trim 결과 비어있지 않은지 검사 후 그렇지 않으면 친절 메시지 후 return.",
     "@sym/ui-cli init 명령",
     "@sym/ui-cli 0.0.0", "@sym/ui-cli 0.1.0", "완료", "Major"),

    (8, "2026-04-27", "bug",
     "CLI runner: pnpm add/npm install 실패 코드 무시 → 설치 실패해도 성공 처리",
     "오프라인 또는 권한 문제로 의존성 설치 실패",
     "1. 네트워크 단절 상태에서 sym-ui add button\n2. pnpm add 가 exit code != 0 반환\n3. CLI 는 ✓ 표시 후 종료\n4. 실제로는 의존성 누락",
     "원인: installPackages 의 결과 code 검사 누락.\n수정: InstallError 클래스 도입, code != 0 시 throw, add/init 에서 catch 후 사용자 친화 메시지로 process.exit. child error 핸들러도 추가.",
     "@sym/ui-cli init/add 명령",
     "@sym/ui-cli 0.0.0", "@sym/ui-cli 0.1.0", "완료", "Major"),

    (9, "2026-04-27", "bug",
     "비핵심 hex 색 혼입 (#a3a6af, #3a3e4c, #ff6b6b)",
     "다크모드 사용",
     "1. accordion.tsx 에 dark:text-[#a3a6af]\n2. radio-group.tsx 에 dark:border-[#3a3e4c]\n3. form.tsx 에 dark:text-[#ff6b6b]",
     "원인: 핵심 팔레트 외 일회성 hex 가 일관성 없이 추가됨.\n수정: 모두 시맨틱 토큰 (muted-foreground, border-input, text-destructive) 으로 치환.",
     "Accordion, RadioGroup, Form",
     "@sym/ui 0.0.0", "@sym/ui 0.1.0", "완료", "Major"),

    (10, "2026-04-27", "bug",
     "Calendar day_disabled / day_range_middle 다크 변형 누락",
     "다크모드에서 날짜 비활성화 또는 범위 선택",
     "1. 다크모드에서 비활성 날짜는 light text-neutral-400 그대로 적용 (충분히 안 어두움)\n2. day_range_middle 도 light bg-neutral-100 그대로",
     "원인: dark: 변형 누락.\n수정: muted-foreground / muted 토큰으로 통일하여 자동 대응.",
     "Calendar, DatePicker",
     "@sym/ui 0.0.0", "@sym/ui 0.1.0", "완료", "Major"),

    (11, "2026-04-27", "bug",
     "Input/Select 다크 포커스 링 색이 primary-900/40 (요구 사양은 ring 토큰)",
     "다크모드에서 Input/Select 포커스",
     "1. 다크모드 Input 포커스 시 ring 색이 너무 어두움 (primary-900/40)\n2. 식별이 어려움",
     "원인: light 포커스 ring (primary-100) 의 단순 dark 대응으로 primary-900/40 채택.\n수정: ring 시맨틱 토큰 (HSL 213 93% 62%) 으로 통일.",
     "Input, Select",
     "@sym/ui 0.0.0", "@sym/ui 0.1.0", "완료", "Major"),

    (12, "2026-04-27", "bug",
     "Card/Input/Select 보더-배경 분리도 부족 (ΔL=5.7)",
     "다크모드에서 카드/입력 보더 식별",
     "1. card 보더 #2a2d3e (L=20.4) on 배경 #1e222d (L=14.7) → ΔL=5.7\n2. 보더가 거의 안 보임",
     "원인: 보더와 배경 L 차이 부족.\n수정: 새 토큰 border (L=28) on surface (L=18) → ΔL=10 으로 확보.",
     "Card, Input, Select, DataTable, Sheet, Dialog 등",
     "@sym/ui 0.0.0", "@sym/ui 0.1.0", "완료", "Major"),

    (13, "2026-04-27", "bug",
     "Tooltip 다크모드 색 반전 (bg-neutral-50 + text-neutral-900)",
     "다크모드에서 Tooltip 표시",
     "1. 다크모드에서 Tooltip 이 흰 배경 검은 글자로 반전 표시\n2. 다크 팔레트 정면 충돌",
     "원인: 라이트와 동일한 정반대 색을 다크에 그대로 적용 (Material 스타일 의도였을 수 있으나 팔레트와 충돌).\n수정: bg-foreground text-background 로 양 모드 자동 반전.",
     "Tooltip",
     "@sym/ui 0.0.0", "@sym/ui 0.1.0", "완료", "Major"),

    (14, "2026-04-27", "bug",
     "Accent 색 혼용 (blue-500 ↔ #3d7eff)",
     "다크모드에서 선택 표시 (RadioGroup, Progress, Calendar)",
     "1. RadioGroup 은 blue-500 (#3b82f6) + dark #3d7eff\n2. Progress 도 같은 혼용\n3. Calendar day_selected 도 같은 혼용",
     "원인: primary 토큰 미정의로 컴포넌트마다 직접 color 지정.\n수정: primary 시맨틱 토큰으로 통일 (light 217 91% 60%, dark 213 93% 62%).",
     "RadioGroup, Progress, Calendar",
     "@sym/ui 0.0.0", "@sym/ui 0.1.0", "완료", "Major"),

    (15, "2026-04-27", "bug",
     "DataTable 정렬 헤더가 click-only <th> (포커스 불가, 키보드 활성 불가, aria-sort 없음)",
     "DataTable enableSorting 사용",
     "1. <th onClick=...> 만 있고 button 없음\n2. Tab 으로 포커스 못 잡음, Enter/Space 동작 안 함\n3. aria-sort 미설정으로 화면낭독기 정렬 상태 인지 불가",
     "원인: 시각 전용 정렬 UI 만 구현.\n수정: th 안에 button 으로 감싸고 aria-sort (ascending/descending/none) 동기화, Enter/Space 핸들링, focus-visible:ring-ring 추가.",
     "DataTable",
     "@sym/ui 0.0.0", "@sym/ui 0.1.0", "완료", "Major"),

    (16, "2026-04-27", "bug",
     "AccordionTrigger 포커스 링 없음 (키보드 사용자 식별 불가)",
     "키보드만 사용해서 Accordion 조작",
     "1. Tab 으로 AccordionTrigger 에 포커스\n2. 시각적 피드백 없음 (오로지 hover:underline 만 존재)",
     "원인: focus-visible 스타일 누락.\n수정: focus-visible:outline-none focus-visible:ring-2 focus-visible:ring-ring focus-visible:ring-offset-2 focus-visible:ring-offset-background rounded-sm 추가.",
     "Accordion",
     "@sym/ui 0.0.0", "@sym/ui 0.1.0", "완료", "Major"),

    (17, "2026-04-27", "bug",
     "Calendar nav/day 버튼 키보드 포커스 시각화 없음",
     "키보드로 Calendar 네비게이션",
     "1. Tab/화살표로 nav 또는 day 에 포커스\n2. hover 만 정의되어 있어 포커스 식별 불가",
     "원인: focus-visible 스타일 누락.\n수정: nav_button, day 모두에 focus-visible:ring-ring + offset-1 추가.",
     "Calendar, DatePicker",
     "@sym/ui 0.0.0", "@sym/ui 0.1.0", "완료", "Major"),

    (18, "2026-04-27", "bug",
     "Combobox CommandInput 이 placeholder 만 있고 프로그래매틱 라벨 없음",
     "Combobox 사용",
     "1. <Combobox /> 렌더\n2. 검색 인풋이 placeholder 만 있음\n3. 화면낭독기 사용자가 인풋 목적 인지 불가 (placeholder 는 라벨이 아님)",
     "원인: aria-label 미설정.\n수정: searchAriaLabel?: string prop 노출, CommandInput 에 aria-label={searchAriaLabel ?? searchPlaceholder} 자동 연결.",
     "Combobox",
     "@sym/ui 0.0.0", "@sym/ui 0.1.0", "완료", "Major"),

    # === Minor 4 ===
    (19, "2026-04-27", "enhancement",
     "Button 에 asChild + Slot 패턴 추가 (shadcn 표준)",
     "",
     "shadcn 표준 패턴인 asChild prop 으로 Link/커스텀 트리거 래핑 가능하게 한다.",
     "사유: shadcn 호환성. Link 안에 Button 스타일을 적용하려면 button 안에 a 태그를 중첩할 수 없으므로 Slot 으로 외부 컴포넌트에 위임해야 함.",
     "Button (다른 컴포넌트는 영향 없음)",
     "@sym/ui 0.0.0", "@sym/ui 0.1.0", "완료", "Minor"),

    (20, "2026-04-27", "bug",
     "Disabled 색이 #787b86 / #4a4d55 로 분산되어 의미 불분명",
     "다크모드에서 disabled 상태 표시",
     "1. Input disabled 는 #787b86\n2. Calendar day_disabled 는 #4a4d55\n3. 같은 의미인데 색이 다름",
     "원인: 일관 규칙 부재.\n수정: 모두 muted-foreground 토큰 + opacity-40 패턴으로 통일.",
     "Input, Calendar 등 disabled 상태가 있는 모든 컴포넌트",
     "@sym/ui 0.0.0", "@sym/ui 0.1.0", "완료", "Minor"),

    (21, "2026-04-27", "bug",
     "보조 텍스트 #787b86 on 다크 배경 대비 4.24:1 (AA 4.5:1 미달)",
     "다크모드에서 본문보다 흐린 보조 텍스트 표시 (Dialog/Sheet description, Calendar head, Tabs inactive)",
     "1. #787b86 (L=49.8) on 다크 배경\n2. 대비 비율 4.24:1\n3. WCAG AA 4.5:1 미달",
     "원인: 다크 보조 텍스트의 절대 밝기 부족.\n수정: muted-foreground 토큰 L=70 으로 상향 (대비 7~8:1, AAA 통과).",
     "Dialog/Sheet description, Calendar head, Tabs, 모든 muted-foreground 사용처",
     "@sym/ui 0.0.0", "@sym/ui 0.1.0", "완료", "Minor"),

    (22, "2026-04-27", "bug",
     "docs lint 9건 (Storybook stories rules-of-hooks 4건 + tailwind.config.cjs no-undef 5건)",
     "pnpm lint 실행",
     "1. pnpm lint 시 docs 패키지에서 9 errors\n2. calendar/combobox/date-picker/toast stories 가 render: () => { React.useState(...) } 패턴 → render 함수가 컴포넌트가 아니라 일반 함수로 처리됨\n3. tailwind.config.cjs 가 module/require no-undef",
     "원인: stories 의 render inline hook 호출 + eslint config 가 .cjs 를 처리하지 않음.\n수정: render inline 을 별도 컴포넌트로 추출, eslint.config.mjs ignores 에 **/*.cjs 추가.",
     "apps/docs (라이브러리에는 영향 없음)",
     "@sym/ui-cli 0.0.0 (도구 측면)", "전 패키지 0.1.0", "완료", "Minor"),

    # === Codex 2차 재검증에서 발견된 회귀 2건 ===
    (23, "2026-04-27", "bug",
     "DataTable 정렬 button 의 onKeyDown 핸들러가 native Enter/Space 와 이중 토글 위험",
     "DataTable enableSorting 사용 + 키보드 Enter/Space 로 정렬",
     "1. 1차 수정에서 정렬 헤더를 button 으로 감싸면서 onKeyDown 도 추가\n2. button 은 native 로 Enter/Space 활성화\n3. onKeyDown 에서 sortHandler 를 한 번 더 호출\n4. 결과: Enter 1회 입력에 정렬 상태가 두 번 바뀔 수 있음",
     "원인: native button 동작에 익숙하지 않은 채 방어적으로 onKeyDown 추가.\n수정: onKeyDown 핸들러 제거, native onClick 만 유지 (Enter/Space 시 button 이 자동으로 onClick 트리거).",
     "DataTable",
     "@sym/ui 0.1.0 (수정 도중)", "@sym/ui 0.1.0", "완료", "Major"),

    (24, "2026-04-27", "bug",
     "Calendar day_range_middle 가 text-foreground 사용 (토큰 통일 누락)",
     "다크모드에서 날짜 범위 선택, 중간 날짜 표시",
     "1. 1차 수정에서 day_range_middle 을 aria-selected:bg-muted aria-selected:text-foreground 로 변경\n2. 의도는 모두 muted-foreground 로 통일이었으나 text-foreground 로 잘못 적용\n3. 의미상 보조 텍스트인데 본문 색을 사용",
     "원인: 마이그레이션 시 토큰 일관성 룰 적용 누락.\n수정: aria-selected:text-foreground → aria-selected:text-muted-foreground 로 일관화.",
     "Calendar, DatePicker (range mode)",
     "@sym/ui 0.1.0 (수정 도중)", "@sym/ui 0.1.0", "완료", "Minor"),

    # === Codex 3차 재검증에서 발견된 토큰 잔존 3건 ===
    (25, "2026-04-27", "bug",
     "Avatar fallback 이 primary-500/700 + text-white 직접 사용 (토큰 우회)",
     "다크모드에서 Avatar fallback 표시",
     "1. avatar.tsx 의 AvatarFallback 이 bg-gradient-to-br from-primary-500 to-primary-700 text-white 사용\n2. 라이트/다크 자동 전환 안 됨 + 새 primary 토큰 우회",
     "원인: 마이그레이션 누락.\n수정: bg-primary text-primary-foreground 로 통일 (그라디언트 제거 = 디자인 단순화).",
     "Avatar",
     "@sym/ui 0.1.0 (수정 도중)", "@sym/ui 0.1.0", "완료", "Minor"),

    (26, "2026-04-27", "bug",
     "Badge primary/danger variant 가 primary-50/700/950, red 등 직접 사용 (토큰 우회)",
     "다크모드에서 Badge primary/danger variant 표시",
     "1. badge.tsx primary variant: bg-primary-50 text-primary-700 dark:bg-primary-950/40 dark:text-primary-300\n2. danger variant: bg-danger/10 text-danger (palette-coupled)",
     "원인: 토큰 마이그레이션 시 색조형 variant 누락.\n수정: primary → bg-primary/15 text-primary, danger → bg-destructive/10 text-destructive 로 단일 토큰 + opacity 패턴.",
     "Badge",
     "@sym/ui 0.1.0 (수정 도중)", "@sym/ui 0.1.0", "완료", "Minor"),

    (27, "2026-04-27", "bug",
     "Dialog/Sheet overlay 가 bg-black/50 하드코딩 (오버레이 토큰 부재)",
     "Dialog/Sheet 열기",
     "1. dialog.tsx, sheet.tsx 의 overlay 가 bg-black/50 하드코딩\n2. 라이트/다크에서 overlay 색조 분리 불가",
     "원인: --overlay 토큰 미정의.\n수정: globals.css 에 --overlay 정의 (light: 222 16% 6%, dark: 0 0% 0%), tailwind.preset.cjs 에 overlay 토큰 추가, dialog/sheet 가 bg-overlay/50 사용.",
     "Dialog, Sheet, globals.css, tailwind.preset.cjs",
     "@sym/ui 0.1.0 (수정 도중)", "@sym/ui 0.1.0", "완료", "Minor"),
]

def build():
    wb = Workbook()
    ws = wb.active
    ws.title = "issues"

    # Header
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(name="Pretendard", color="FFFFFF", bold=True)
    for col_idx, h in enumerate(HEADERS, start=1):
        c = ws.cell(row=1, column=col_idx, value=h)
        c.fill = header_fill
        c.font = header_font
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Body
    body_font = Font(name="Pretendard")
    wrap_align = Alignment(vertical="top", wrap_text=True)
    for row_idx, row in enumerate(ROWS, start=2):
        for col_idx, val in enumerate(row, start=1):
            c = ws.cell(row=row_idx, column=col_idx, value=val)
            c.font = body_font
            c.alignment = wrap_align

    # Column widths
    widths = [8, 12, 12, 50, 24, 50, 60, 30, 22, 22, 8, 10]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # Freeze header
    ws.freeze_panes = "A2"

    OUT.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUT)
    print(f"issue_log.xlsx written: {OUT} ({len(ROWS)} rows)")

if __name__ == "__main__":
    build()
